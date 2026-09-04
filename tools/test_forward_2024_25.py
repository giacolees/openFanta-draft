#!/usr/bin/env python3.10
"""
TEST FORWARD del modello — statistiche 2024/25 -> fantamedia 2025/26.

Stesso identico modello di build_stats_rating_2025_26.py (stesse feature per ruolo,
stessa ridge pesata per ruolo, stessi parametri, LOOCV), ma allenato sulle
statistiche della stagione 2024/2025 e testato contro la fantamedia 25/26
(e il FVM 26/27) presi dall'xlsx ufficiale.

Cosa misura:
- generalizzazione cross-sezionale: i pesi stimati su un anno spiegano l'anno dopo;
- qui niente e' "in-sample": le feature sono della stagione PRECEDENTE rispetto al target,
  quindi le metriche dicono quanto le statistiche di un anno predicono la FM dell'anno dopo
  (la vera domanda di un fantallenatore: i numeri dell'anno scorso valgono quest'anno?).

Output:
- data/Forward_Test_2024_25_on_FM_2025_26.csv   (per giocatore: rating, PredFM, FM 25/26, FVM)
- data/audit_forward_2024_25.csv / .txt          (metriche per ruolo + importanza feature)

Uso:  python3 tools/test_forward_2024_25.py
"""

import csv
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRAPED_CSV = os.path.join(
    BASE_DIR, "data", "serie_a_stats_2024_25", "giocatori_serie_a_2024_25.csv"
)
QUOTAZIONI_XLSX = os.path.join(
    BASE_DIR, "data", "Quotazioni_Fantacalcio_Stagione_2026_27_arricchito.xlsx"
)
OUT_CSV = os.path.join(BASE_DIR, "data", "Forward_Test_2024_25_on_FM_2025_26.csv")
OUT_AUDIT = os.path.join(BASE_DIR, "data", "audit_forward_2024_25.csv")
OUT_TXT = os.path.join(BASE_DIR, "data", "audit_forward_2024_25.txt")

# Importa il modulo del rating come libreria (main() e' protetto dal guard __main__)
sys.path.insert(0, os.path.join(BASE_DIR, "tools"))
import build_stats_rating_2025_26 as base

BASE = base  # alias per il resto del file

base.SCRAPED_CSV = SCRAPED_CSV
base.QUOTAZIONI_XLSX = QUOTAZIONI_XLSX

KNOWN_SDP_ROLES = {"Goalkeeper", "Defender", "Midfielder", "Forward"}
OUT_COLUMNS = [
    "Rating",
    "PredFM_25_26",
    "FM_25_26",
    "FVM",
    "Ruolo",
    "Nome",
    "Squadra_24_25",
    "Pres_24_25",
    "Min_24_25",
    "Dati",
]


def pname(row):
    """Nome leggibile: nella 24/25 il displayName e' spesso vuoto, uso il nome completo."""
    return (row.get("Nome") or "").strip() or (row.get("NomeCompleto") or "").strip()


def load_quotazioni_with_fvm():
    """Righe xlsx con ruolo, FM 25/26 e FVM (stesso formato di base.load_quotazioni)."""
    wb = openpyxl.load_workbook(QUOTAZIONI_XLSX, data_only=True)
    players = []
    for sheet, role in BASE.XLSX_SHEETS:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[1]]
        idx = {name: i for i, name in enumerate(header) if name}
        for row in rows[2:]:
            if not row or not row[idx["Nome"]]:
                continue
            players.append(
                {
                    "R": role,
                    "Nome": str(row[idx["Nome"]]).strip(),
                    "Squadra": str(row[idx["Squadra"]] or "").strip(),
                    "FM": BASE.to_float(row[idx["FM 25/26"]]),
                    "FVM": BASE.to_float(row[idx["FVM"]]),
                }
            )
    return players


def main():
    scraped = BASE.load_scraped()
    quotazioni = load_quotazioni_with_fvm()
    index = BASE.build_scraped_index(scraped)

    # ruolo: quello dell'xlsx per i matchati (autoritativo), la roleLabel SDP per gli altri
    role_of = {
        r["Id"]: (
            BASE.SDP_TO_FANTASY.get(r["Ruolo"])
            if r["Ruolo"] in KNOWN_SDP_ROLES
            else None
        )
        for r in scraped
    }
    matched_fm = {}
    matched_fvm = {}
    unmatched = []
    for x in quotazioni:
        if x["FM"] is None or x["FM"] <= 0:
            continue
        sp = BASE.match_player(
            x["Nome"], BASE.FANTASY_TO_SDP[x["R"]], x["Squadra"], index
        )
        if sp is None:
            unmatched.append(x["Nome"])
            continue
        pid = sp["row"]["Id"]
        matched_fm[pid] = x["FM"]
        matched_fvm[pid] = x["FVM"]
        role_of[pid] = x["R"]
    print(
        f"Giocatori 24/25: {len(scraped)} | FM 25/26 matchati: {len(matched_fm)} | persi: {len(unmatched)}"
    )

    # setta i globali del modulo base e allena lo STESSO modello su 24/25
    BASE.SCRAPED = scraped
    BASE.MATCHED = matched_fm
    BASE.ROLE_OF = role_of
    models = {role: BASE.fit_role_model(role) for role in ("P", "D", "C", "A")}
    ratings = BASE.compute_ratings(models)

    # ---- output per giocatore ----
    rows_out = []
    by_role = {}
    for r in ratings:
        by_role.setdefault(r["role"], []).append(r)
    for role, group in by_role.items():
        group.sort(key=lambda r: r["rating"], reverse=True)
        for r in group:
            rows_out.append(
                {
                    "Rating": round(r["rating"], 1),
                    "PredFM_25_26": round(r["pred"], 2),
                    "FM_25_26": round(r["fm"], 2) if r["has_fm"] else None,
                    "FVM": round(matched_fvm.get(r["row"]["Id"]) or 0, 0)
                    if r["row"]["Id"] in matched_fvm
                    else None,
                    "Ruolo": BASE.ROLE_FANTASY[role],
                    "Nome": pname(r["row"]),
                    "Squadra_24_25": r["row"]["Squadra"],
                    "Pres_24_25": BASE.safe_int(r["row"].get("Pres")),
                    "Min_24_25": BASE.safe_int(r["minutes"]),
                    "Dati": "backtest FM 25/26" if r["has_fm"] else "senza FM xlsx",
                }
            )
    rows_out.sort(key=lambda r: (r["Ruolo"], -r["Rating"]))

    try:
        with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=OUT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows_out)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_CSV}: {err}") from err
    print(f"CSV: {os.path.relpath(OUT_CSV, BASE_DIR)} ({len(rows_out)} giocatori)")

    # ---- audit ----
    run_audit(models, ratings, matched_fvm, unmatched)
    write_xlsx(rows_out)


def write_xlsx(rows):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    sheets = [("Tutti", rows)] + [
        (
            BASE.ROLE_FANTASY[role],
            [r for r in rows if r["Ruolo"] == BASE.ROLE_FANTASY[role]],
        )
        for role in ("P", "D", "C", "A")
    ]
    path = OUT_CSV.replace(".csv", ".xlsx")
    for name, sheet_rows in sheets:
        ws = wb.create_sheet(name)
        ws.append(OUT_COLUMNS)
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center")
        for r in sheet_rows:
            ws.append([r[c] for c in OUT_COLUMNS])
        for col_idx, col in enumerate(OUT_COLUMNS, start=1):
            width = max(len(col) + 2, 8) if col != "Nome" else 20
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
    del wb["Sheet"]
    try:
        wb.save(path)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {path}: {err}") from err


def run_audit(models, ratings, matched_fvm, unmatched_list):
    lines = []
    audit_rows = []
    lines.append("=" * 80)
    lines.append("AUDIT FORWARD — modello (feature 24/25) -> FM 25/26 / FVM")
    lines.append("=" * 80)
    lines.append(
        f"\nJoin xlsx->scraped 24/25: {sum(1 for r in ratings if r['has_fm'])} giocatori con FM 25/26.  "
        f"Non matchati: {unmatched_list}"
    )
    lines.append(
        "Nota: feature della stagione 24/25, target della 25/26: nessun dato e' in-sample."
    )

    for role in ("P", "D", "C", "A"):
        m = models[role]
        cv = m["cv"]
        group = [r for r in ratings if r["role"] == role and r["has_fm"]]
        fvm_list = [
            (r["pred"], matched_fvm[r["row"]["Id"]])
            for r in group
            if r["row"]["Id"] in matched_fvm and matched_fvm[r["row"]["Id"]]
        ]
        pe_fvm = (
            BASE.pearson([p for p, _ in fvm_list], [v for _, v in fvm_list])
            if len(fvm_list) > 3
            else 0.0
        )
        sp_fvm = (
            BASE.spearman([p for p, _ in fvm_list], [v for _, v in fvm_list])
            if len(fvm_list) > 3
            else 0.0
        )
        lines.append(
            f"\n--- {BASE.ROLE_FANTASY[role]} ({role}) — n train={m['n_train']}, tot={m['n_total']}, "
            f"con FM 25/26: {len(group)} ---"
        )
        lines.append(
            f"  LOOCV (24/25->FM 25/26) : R2={cv['r2']:.3f}  Pearson={cv['pearson']:.3f}  "
            f"Spearman={cv['spearman']:.3f}  RMSE={cv['rmse']:.3f} (FM)"
        )
        lines.append(
            f"  PredFM(24/25) vs FVM     : Pearson={pe_fvm:.3f}  Spearman={sp_fvm:.3f}  (n={len(fvm_list)})"
        )
        lines.append("  feature importance (effetto FM per +1 dev std):")
        for feat, b in sorted(m["importance"], key=lambda t: -abs(t[1])):
            lines.append(f"    {feat[0]:22} beta={b:+.3f}")
        audit_rows.append(
            {
                "ruolo": role,
                "n_train": m["n_train"],
                "n_total": m["n_total"],
                "n_fm": len(group),
                "cv_r2": round(cv["r2"], 3),
                "cv_pearson": round(cv["pearson"], 3),
                "cv_spearman": round(cv["spearman"], 3),
                "cv_rmse": round(cv["rmse"], 3),
                "fvm_pearson": round(pe_fvm, 3),
                "fvm_spearman": round(sp_fvm, 3),
            }
        )
    lines.append("\n" + "=" * 80)

    lines.append(
        "\nRESIDUI piu' forti (FM 25/26 reale - PredFM da 24/25, solo con FM):"
    )
    for role in ("P", "D", "C", "A"):
        group = [
            r
            for r in ratings
            if r["role"] == role and r["has_fm"] and r["minutes"] >= BASE.MIN_MINUTES
        ]
        group.sort(key=lambda r: r["fm"] - r["pred"])
        worst = group[:3] + group[-3:]
        tag = ["sottostimati"] * 3 + ["sovrastimati"] * 3
        lines.append(f"  {BASE.ROLE_FANTASY[role]}:")
        for r, t in zip(worst, tag, strict=True):
            lines.append(
                f"    {t:12} {pname(r['row']):20} {r['row']['Squadra']:12} "
                f"FM={r['fm']:.2f} pred={r['pred']:.2f} d={r['fm'] - r['pred']:+.2f}"
            )

    lines.append("\nOVERLAP TOP-10 (rating 24/25 vs FM 25/26 reale e vs FVM):")
    for role in ("P", "D", "C", "A"):
        group = [r for r in ratings if r["role"] == role and r["has_fm"]]
        by_fm = sorted(group, key=lambda r: -r["fm"])[:10]
        by_rating = sorted(group, key=lambda r: -r["rating"])[:10]
        fm_ids = {r["row"]["Id"] for r in by_fm}
        hit_fm = sum(1 for r in by_rating if r["row"]["Id"] in fm_ids)
        fvm_group = [
            (r["row"]["Id"], matched_fvm[r["row"]["Id"]])
            for r in group
            if r["row"]["Id"] in matched_fvm and matched_fvm[r["row"]["Id"]]
        ]
        top_fvm = {pid for pid, _ in sorted(fvm_group, key=lambda z: -z[1])[:10]}
        hit_fvm = (
            sum(1 for r in by_rating if r["row"]["Id"] in top_fvm) if top_fvm else 0
        )
        lines.append(
            f"  {BASE.ROLE_FANTASY[role]}: vs FM 25/26 {hit_fm}/10 | vs FVM {hit_fvm}/10"
        )
        audit_rows[-1]["overlap_fm_top10"] = f"{hit_fm}/10"
        audit_rows[-1]["overlap_fvm_top10"] = f"{hit_fvm}/10"

    with_fm = sum(1 for r in ratings if r["has_fm"])
    lines.append(
        f"\nRating calcolati: {len(ratings)} giocatori (con FM 25/26 per backtest: {with_fm})"
    )

    try:
        with open(OUT_AUDIT, "w", newline="", encoding="utf-8-sig") as f:
            cols = [
                "ruolo",
                "n_train",
                "n_total",
                "n_fm",
                "cv_r2",
                "cv_pearson",
                "cv_spearman",
                "cv_rmse",
                "fvm_pearson",
                "fvm_spearman",
                "overlap_fm_top10",
                "overlap_fvm_top10",
            ]
            writer = csv.DictWriter(f, fieldnames=cols)
            writer.writeheader()
            writer.writerows(audit_rows)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_AUDIT}: {err}") from err

    print("\n".join(lines))
    try:
        with open(OUT_TXT, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_TXT}: {err}") from err


if __name__ == "__main__":
    main()
