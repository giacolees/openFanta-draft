#!/usr/bin/env python3.10
"""
Rating statistico Serie A 2025/26 (scala 0-100 per ruolo) — versione data-driven.

Idea: invece di pesi fissati a mano, i pesi delle feature statistiche vengono
STIMATI dai dati, usando come target la fantamedia ufficiale FM 25/26
(data/Quotazioni_Fantacalcio_Stagione_2026_27_arricchito.xlsx) per il backtest.

Pipeline:
1. Join giocatori scrapati (data/serie_a_stats_2025_26/giocatori_serie_a_2025_26.csv)
   con il ruolo/la fantamedia dell'xlsx (match per nome+squadra+ruolo+iniziale, fuzzy).
2. Feature engineering per ruolo (P/D/C/A): metriche selezionate a mano tra le ~350
   disponibili, con rate per 90 minuti e percentuali. Feature scarne o piatte
   (quasi tutte a zero) vengono scartate.
3. Ridge regression pesata per ruolo (minuti bassi => peso minore), risolta senza
   numpy (equazioni normali + eliminazione gaussiana). Pesature standardizzate.
4. Cross-validazione 5-fold per ruolo: R2, Pearson, Spearman, RMSE su FM.
5. Rating finale 0-100 = percentile dell'FM predetto dentro il ruolo, con shrinkage
   sui minuti (pochi minuti => verso la mediana del ruolo). Tier S/A/B/C/D.
6. Audit: coverage del join, metriche CV per ruolo, feature importance, top
   giocatori sotto/sopravvalutati, overlap con i Top-10 FM reali.

Output:
- data/Stats_Rating_Stagione_2025_26.csv / .xlsx
- data/audit_stats_rating_2025_26.csv  (+ report a console)

Uso:  python3 scripts/build_stats_rating_2025_26.py
"""

import csv
import os
import re
import statistics
import unicodedata
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRAPED_CSV = os.path.join(
    BASE_DIR, "data", "serie_a_stats_2025_26", "giocatori_serie_a_2025_26.csv"
)
QUOTAZIONI_XLSX = os.path.join(
    BASE_DIR, "data", "Quotazioni_Fantacalcio_Stagione_2026_27_arricchito.xlsx"
)
OUT_CSV = os.path.join(BASE_DIR, "data", "Stats_Rating_Stagione_2025_26.csv")
OUT_XLSX = os.path.join(BASE_DIR, "data", "Stats_Rating_Stagione_2025_26.xlsx")
OUT_AUDIT = os.path.join(BASE_DIR, "data", "audit_stats_rating_2025_26.csv")
OUT_TXT = os.path.join(BASE_DIR, "data", "audit_stats_rating_2025_26.txt")

ROLE_FANTASY = {
    "P": "Portiere",
    "D": "Difensore",
    "C": "Centrocampista",
    "A": "Attaccante",
}
FANTASY_TO_SDP = {"P": "Goalkeeper", "D": "Defender", "C": "Midfielder", "A": "Forward"}
SDP_TO_FANTASY = {v: k for k, v in FANTASY_TO_SDP.items()}

MIN_MINUTES = 270  # sotto questa soglia i rate /90 sono rumorosi (peso ridotto,
# in esclusi dal training per la stima dei pesi)
# regolarizzazione ridge per ruolo (X standardizzate): piu' alto con n piccolo
LAMBDA_BY_ROLE = {"P": 3.0, "D": 1.0, "C": 1.0, "A": 1.5}

# Feature per ruolo: (colonna CSV, come usarla)
#   "per90"  -> valore * 90 / minuti
#   "perc"   -> valore gia' in percentuale
#   "ratio"  -> valore / presenze (es. clean sheet per partita)
#   "raw"    -> valore cosi' com'e'
# Sono state scartate le feature piatte (es. Aerials-Won-Perc, xGEfficienza, ...)
# e quelle ridondanti rispetto a metriche piu' informative (es. shotsSaved=0 per tutti).
FEATURES = {
    "P": [
        ("Saves Made", "per90"),  # parate
        ("totalSavePerc", "perc"),  # % parate (segnale principale)
        ("Clean sheets", "ratio"),  # puliti per partita
        ("Goals Conceded", "per90"),  # gol subiti (coeff. atteso negativo)
        ("Saves from Penalty", "per90"),  # rigori parati
    ],
    "D": [
        ("Gol", "per90"),
        ("Assist", "per90"),
        ("Clean sheets", "ratio"),
        ("TackleVinti", "per90"),
        ("Interceptions", "per90"),
        ("Total Clearances", "per90"),
        ("Blocked Shots", "per90"),
        ("DuelliVintiPerc", "perc"),
        ("PassPrec", "perc"),
        ("xG", "per90"),
        ("PassChiave", "per90"),
        ("Successful Dribbles", "per90"),
    ],
    "C": [
        ("Gol", "per90"),
        ("Assist", "per90"),
        ("xG", "per90"),
        ("TiriPorta", "per90"),
        ("PassChiave", "per90"),
        ("PassPrec", "perc"),
        ("DuelliVintiPerc", "perc"),
        ("TackleVinti", "per90"),
        ("Interceptions", "per90"),
        ("Successful Dribbles", "per90"),
    ],
    "A": [
        ("Gol", "per90"),
        ("Assist", "per90"),
        ("xG", "per90"),
        ("Tiri", "per90"),
        ("TiriPorta", "per90"),
        ("PassChiave", "per90"),
        ("bigChance", "per90"),
        ("bigChanceConvert", "per90"),
        ("Successful Dribbles", "per90"),
    ],
}

XLSX_SHEETS = [
    ("Portieri", "P"),
    ("Difensori", "D"),
    ("Centrocampisti", "C"),
    ("Attaccanti", "A"),
]
OUT_COLUMNS = [
    "Rating",
    "Tier",
    "Rank ruolo",
    "Ruolo",
    "Nome",
    "Squadra",
    "Id",
    "Pres",
    "Min",
    "FM 25/26",
    "PredFM",
    "Dati",
]

# --------------------------------------------------------------------------- #
# normalizzazione nomi e matching fuzzy
# --------------------------------------------------------------------------- #

TRANS = {
    "ø": "o",
    "Ø": "O",
    "æ": "ae",
    "Æ": "AE",
    "ß": "ss",
    "ı": "i",
    "đ": "d",
    "Đ": "D",
    "ł": "l",
    "Ł": "L",
    "ç": "c",
    "ğ": "g",
    "ş": "s",
    "ñ": "n",
}
STOP = {"de", "del", "della", "van", "di", "da", "dos", "das", "o", "e", "y"}


def norm_name(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    for a, b in TRANS.items():
        s = s.replace(a, b)
    s = s.lower()
    s = re.sub(r"[^a-z0-9' ]", " ", s)
    return s.replace("'", " ")


def name_tokens(*parts):
    """Token significativi da uno o piu' campi nome (display + nome completo)."""
    base = [
        t for p in parts for t in norm_name(p).split() if len(t) > 1 and t not in STOP
    ]
    out = []
    for t in base:
        no_ap = t.replace("'", "")
        if no_ap not in out:
            out.append(no_ap)
    return out


def initial_of(s):
    m = re.search(r"\b([a-z0-9])\.", (s or "").lower())
    return m.group(1) if m else None


def match_player(x_name, x_role, x_team, scraped_index):
    """
    Trova il giocatore scrapato corrispondente a una riga xlsx.
    Ritorna il dict scraped, oppure None.
    """
    txs = set(name_tokens(x_name))
    if not txs:
        return None
    xinit = initial_of(x_name)
    scored = []
    for sp in scraped_index:
        sts = sp["tokens"]
        inter = 0
        for t in txs:
            if t in sts:
                inter += 1
            elif len(t) >= 4 and any(st.startswith(t) for st in sts):
                inter += 1  # contenimento (es. 'dicka' in 'ndicka')
        if inter == 0:
            continue
        significant = any(
            len(t) >= 3 and (t in sts or any(st.startswith(t) for st in sts))
            for t in txs
        )
        if not significant and inter < len(txs):
            continue
        score = inter * 2
        if sp["row"]["Squadra"] == x_team:
            score += 2
        if sp["row"]["Ruolo"] == x_role:
            score += 1
        if xinit and sp["init"] == xinit:
            score += 2  # iniziale: segnale forte (gemelli omonimi)
        scored.append((score, sp))
    if not scored:
        return None
    scored.sort(key=lambda z: -z[0])
    top = scored[0][0]
    bests = [sp for s, sp in scored if s == top]
    return bests[0] if len(bests) == 1 else None


# --------------------------------------------------------------------------- #
# utilita' numeriche (pure python, niente numpy)
# --------------------------------------------------------------------------- #


def to_float(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return 0.0 + value  # int/float -> float, mai un errore
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def safe_int(value, default=0):
    try:
        v = to_float(value)
        return int(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def pearson(a, b):
    n = len(a)
    if n < 2:
        return 0.0
    ma, mb = sum(a) / n, sum(b) / n
    cov = sum((x - ma) * (y - mb) for x, y in zip(a, b, strict=True))
    va = sum((x - ma) ** 2 for x in a)
    vb = sum((y - mb) ** 2 for y in b)
    if va == 0 or vb == 0:
        return 0.0
    return cov / (va**0.5 * vb**0.5)


def spearman(a, b):
    def ranks(v):
        order = sorted(range(len(v)), key=lambda i: v[i])
        r = [0.0] * len(v)
        i = 0
        while i < len(order):
            j = i
            while j + 1 < len(order) and v[order[j + 1]] == v[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r

    return pearson(ranks(a), ranks(b))


def solve_linear(a, b):
    """Eliminazione gaussiana con pivot parziale (a: n x n, b: n)."""
    n = len(b)
    a = [row[:] for row in a]
    b = b[:]
    for i in range(n):
        piv = max(range(i, n), key=lambda r: abs(a[r][i]))
        if abs(a[piv][i]) < 1e-12:
            raise ValueError("matrice singolare")
        a[i], a[piv] = a[piv], a[i]
        b[i], b[piv] = b[piv], b[i]
        for r in range(i + 1, n):
            f = a[r][i] / a[i][i]
            for c in range(i, n):
                a[r][c] -= f * a[i][c]
            b[r] -= f * b[i]
    x = [0.0] * n
    for i in range(n - 1, -1, -1):
        x[i] = (b[i] - sum(a[i][j] * x[j] for j in range(i + 1, n))) / a[i][i]
    return x


def weighted_ridge(X, y, w, lam):
    """Ridge pesata via equazioni normali. X: lista righe (gia' standardizzate)."""
    n_feat = len(X[0])
    xtwx = [[0.0] * n_feat for _ in range(n_feat)]
    xtwy = [0.0] * n_feat
    for row, yi, wi in zip(X, y, w, strict=True):
        for i in range(n_feat):
            xtwy[i] += wi * row[i] * yi
            for j in range(i, n_feat):
                xtwx[i][j] += wi * row[i] * row[j]
    for i in range(n_feat):
        for j in range(i):
            xtwx[i][j] = xtwx[j][i]
        xtwx[i][i] += lam
    return solve_linear(xtwx, xtwy)


def standardize(matrix):
    """Standardizza ogni colonna (media 0, dev std 1). Ritorna (Z, mean, std)."""
    n_cols = len(matrix[0])
    mean, std = [], []
    for c in range(n_cols):
        col = [r[c] for r in matrix]
        m = sum(col) / len(col)
        sd = (sum((v - m) ** 2 for v in col) / len(col)) ** 0.5
        if sd == 0:
            sd = 1.0
        mean.append(m)
        std.append(sd)
    z = [[(r[c] - mean[c]) / std[c] for c in range(n_cols)] for r in matrix]
    return z, mean, std


def percentile_of(value, sorted_values):
    """Quantile (0-1) di `value` dentro una lista ordinata."""
    if not sorted_values:
        return 0.5
    below = sum(1 for v in sorted_values if v < value)
    equal = sum(1 for v in sorted_values if v == value)
    return (below + equal / 2) / len(sorted_values)


def tier_of(fraction):
    if fraction >= 0.90:
        return "S"
    if fraction >= 0.70:
        return "A"
    if fraction >= 0.30:
        return "B"
    if fraction >= 0.10:
        return "C"
    return "D"


# --------------------------------------------------------------------------- #
# caricamento dati
# --------------------------------------------------------------------------- #


def load_scraped():
    try:
        with open(SCRAPED_CSV, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    except OSError as err:
        raise RuntimeError(f"Impossibile leggere {SCRAPED_CSV}: {err}") from err
    seen = set()
    out = []
    for r in rows:
        if not (r.get("NomeCompleto") or r.get("Nome") or "").strip():
            continue
        if r["Id"] in seen:
            continue  # duplicati con lo stesso Id restituiti dall'API
        seen.add(r["Id"])
        out.append(r)
    return out


def load_quotazioni():
    wb = openpyxl.load_workbook(QUOTAZIONI_XLSX, data_only=True)
    players = []
    for sheet, role in XLSX_SHEETS:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[1]]
        idx = {name: i for i, name in enumerate(header) if name}
        for row in rows[2:]:
            if not row or not row[idx["Nome"]]:
                continue
            players.append(
                {
                    "R": role,
                    "Nome": str(row[idx["Nome"]]).strip(),
                    "Squadra": str(row[idx["Squadra"]] or "").strip(),
                    "FM": to_float(row[idx["FM 25/26"]]),
                }
            )
    return players


def build_scraped_index(scraped):
    index = []
    for r in scraped:
        index.append(
            {
                "row": r,
                "tokens": name_tokens(r["NomeCompleto"], r["Nome"]),
                "init": initial_of(r["NomeCompleto"]) or initial_of(r["Nome"]),
            }
        )
    return index


# --------------------------------------------------------------------------- #
# feature engineering
# --------------------------------------------------------------------------- #


def minutes_of(row):
    m = to_float(row.get("Min"))
    if m:
        return m
    pres = to_float(row.get("Pres"))
    return pres * 90 if pres else 0.0


def feature_vector(row, features, minutes):
    vec = []
    for col, mode in features:
        v = to_float(row.get(col))
        if v is None:
            vec.append(None)
            continue
        if mode == "per90":
            vec.append(v * 90 / minutes if minutes > 0 else None)
        elif mode == "ratio":
            pres = to_float(row.get("Pres"))
            vec.append(v / pres if pres else None)
        else:
            vec.append(v)
    return vec


def build_role_dataset(role, scraped, matched, role_of):
    """
    Ritorna (rows, train) per un ruolo.
      - rows:  TUTTI i giocatori scrapati del ruolo (features imputate con la mediana di ruolo)
      - train: sottinsieme con FM > 0 (per fit e CV)
    """
    features = FEATURES[role]
    rows = []
    for sp in scraped:
        if role_of.get(sp["Id"]) != role:
            continue
        minutes = minutes_of(sp)
        vec = feature_vector(sp, features, minutes)
        rows.append(
            {"row": sp, "fm": matched.get(sp["Id"]), "minutes": minutes, "vec": vec}
        )
    # imputazione: le colonne mancanti (es. Clean sheets per chi non ne ha) prendono
    # la mediana del ruolo, cosi' il giocatore resta nel dataset invece di sparire
    n_feat = len(features)
    for c in range(n_feat):
        vals = sorted(r["vec"][c] for r in rows if r["vec"][c] is not None)
        med = statistics.median(vals) if vals else 0.0
        for r in rows:
            if r["vec"][c] is None:
                r["vec"][c] = med
        # winsorizzazione al 1°/99° percentile del ruolo: le code estreme
        # (es. parate/90 di un portiere con poche partite) sballano la regressione
        n = len(vals)
        if n >= 10:
            p1 = vals[n // 100]
            p99 = vals[min(n - 1, 99 * n // 100)]
            for r in rows:
                r["vec"][c] = max(p1, min(p99, r["vec"][c]))
    train = [
        r
        for r in rows
        if r["fm"] is not None and r["fm"] > 0 and r["minutes"] >= MIN_MINUTES
    ]
    return rows, train


# --------------------------------------------------------------------------- #
# modello (ridge per ruolo) + CV
# --------------------------------------------------------------------------- #


def fit_role_model(role):
    """Ritorna il modello (mean/std/beta), le metriche CV e le feature importance."""
    features = FEATURES[role]
    rows, train = build_role_dataset(role, SCRAPED, MATCHED, ROLE_OF)
    n_feat = len(features)

    X = [r["vec"] for r in train]
    y = [r["fm"] for r in train]
    w = [
        min(1.0, (r["minutes"] / MIN_MINUTES) ** 0.5) if r["minutes"] else 0.0
        for r in train
    ]

    # ---- CV leave-one-out (out-of-sample): deterministica e adatta ai n piccoli ----
    n_train = len(train)
    pred_all, y_all = [], []
    for i in range(n_train):
        tr = train[:i] + train[i + 1 :]
        te = train[i]
        Xtr = [r["vec"] for r in tr]
        ytr = [r["fm"] for r in tr]
        wtr = [
            min(1.0, (r["minutes"] / MIN_MINUTES) ** 0.5) if r["minutes"] else 0.0
            for r in tr
        ]
        Ztr, mean, std = standardize(Xtr)
        ybar_tr = (
            sum(w * yv for w, yv in zip(wtr, ytr, strict=True)) / sum(wtr)
            if wtr
            else 0.0
        )
        beta = weighted_ridge(
            Ztr, [yv - ybar_tr for yv in ytr], wtr, LAMBDA_BY_ROLE[role]
        )
        z = [(te["vec"][c] - mean[c]) / std[c] for c in range(n_feat)]
        pred = ybar_tr + sum(b * zz for b, zz in zip(beta, z, strict=True))
        pred_all.append(pred)
        y_all.append(te["fm"])

    def metrics(pred, actual):
        n = len(pred)
        mean_y = sum(actual) / n
        ss_tot = sum((a - mean_y) ** 2 for a in actual)
        ss_res = sum((p - a) ** 2 for p, a in zip(pred, actual, strict=True))
        r2 = 1 - ss_res / ss_tot if ss_tot else 0.0
        rmse = (ss_res / n) ** 0.5
        return {
            "r2": r2,
            "rmse": rmse,
            "pearson": pearson(pred, actual),
            "spearman": spearman(pred, actual),
        }

    cv = metrics(pred_all, y_all)

    # ---- fit finale su tutto il training set ----
    Z, mean, std = standardize(X)
    ybar = sum(w * yv for w, yv in zip(w, y, strict=True)) / sum(w) if w else 0.0
    beta = weighted_ridge(Z, [yv - ybar for yv in y], w, LAMBDA_BY_ROLE[role])
    # importance: effetto FM per +1 dev std della feature (beta e' su X standardizzate)
    importance = [(feat, b) for feat, b in zip(features, beta, strict=True)]
    # banda FM osservata nel training: le predizioni extra-campo vengono clampate
    fm_sorted = sorted(y)
    n = len(fm_sorted)
    fm_lo = fm_sorted[n // 20] if n else 4.0
    fm_hi = fm_sorted[min(n - 1, 19 * n // 20)] if n else 8.0

    return {
        "role": role,
        "mean": mean,
        "std": std,
        "beta": beta,
        "ybar": ybar,
        "fm_lo": fm_lo,
        "fm_hi": fm_hi,
        "cv": cv,
        "n_train": len(train),
        "n_total": len(rows),
        "importance": importance,
        "features": features,
    }


def predict(model, vec):
    z = [(vec[c] - model["mean"][c]) / model["std"][c] for c in range(len(vec))]
    pred = model["ybar"] + sum(b * zz for b, zz in zip(model["beta"], z, strict=True))
    return max(
        model["fm_lo"], min(model["fm_hi"], pred)
    )  # niente estrapolazioni assurde


def compute_ratings(models):
    """Rating 0-100 per ruolo (shrinkage su minuti) per TUTTI i giocatori scrapati."""
    out = []
    for role, model in models.items():
        rows, _ = build_role_dataset(role, SCRAPED, MATCHED, ROLE_OF)
        preds = [(r, predict(model, r["vec"])) for r in rows]
        role_mean = sum(p for _, p in preds) / len(preds) if preds else 6.5
        shrunk = []
        for r, p in preds:
            rel = min(1.0, (r["minutes"] / MIN_MINUTES) ** 0.5) if r["minutes"] else 0.0
            shrunk.append(rel * p + (1 - rel) * role_mean)
        ordered = sorted(shrunk)
        for (r, p), s in zip(preds, shrunk, strict=True):
            rating = 100 * percentile_of(s, ordered)
            out.append(
                {
                    "rating": rating,
                    "pred": p,
                    "shrunk": s,
                    "role": role,
                    "row": r["row"],
                    "fm": r["fm"],
                    "minutes": r["minutes"],
                    "has_fm": r["fm"] is not None and r["fm"] > 0,
                }
            )
    return out


# --------------------------------------------------------------------------- #
# output
# --------------------------------------------------------------------------- #


def write_csv(rows):
    try:
        with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=OUT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_CSV}: {err}") from err


def write_xlsx(rows):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    sheets = [("Tutti", rows)] + [
        (ROLE_FANTASY[role], [r for r in rows if r["Ruolo"] == ROLE_FANTASY[role]])
        for role in ("P", "D", "C", "A")
    ]
    for name, sheet_rows in sheets:
        ws = wb.create_sheet(name)
        ws.append(OUT_COLUMNS)
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center")
        for r in sheet_rows:
            ws.append([r[c] for c in OUT_COLUMNS])
        for col_idx, col in enumerate(OUT_COLUMNS, start=1):
            width = max(len(col) + 2, 8) if col != "Nome" else 20
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
    del wb["Sheet"]
    try:
        wb.save(OUT_XLSX)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_XLSX}: {err}") from err


def tier(rank_fraction):
    if rank_fraction >= 0.90:
        return "S"
    if rank_fraction >= 0.70:
        return "A"
    if rank_fraction >= 0.30:
        return "B"
    if rank_fraction >= 0.10:
        return "C"
    return "D"


def build_rows(ratings):
    by_role = defaultdict(list)
    for r in ratings:
        by_role[r["role"]].append(r)
    out = []
    for role, group in by_role.items():
        group.sort(key=lambda r: r["rating"], reverse=True)
        n = len(group)
        for i, r in enumerate(group, start=1):
            out.append(
                {
                    "Rating": round(r["rating"], 1),
                    "Tier": tier(1 - (i - 1) / max(n - 1, 1)),
                    "Rank ruolo": i,
                    "Ruolo": ROLE_FANTASY[role],
                    "Nome": r["row"]["Nome"],
                    "Squadra": r["row"]["Squadra"],
                    "Id": r["row"]["Id"],
                    "Pres": safe_int(r["row"].get("Pres")),
                    "Min": safe_int(r["minutes"]),
                    "FM 25/26": round(r["fm"], 2) if r["has_fm"] else None,
                    "PredFM": round(r["pred"], 2),
                    "Dati": "25/26 (backtest FM)"
                    if r["has_fm"]
                    else "25/26 (senza FM xlsx)",
                }
            )
    out.sort(key=lambda r: (r["Ruolo"], r["Rank ruolo"]))
    return out


# --------------------------------------------------------------------------- #
# audit
# --------------------------------------------------------------------------- #


def run_audit(models, ratings, matched_count, unmatched_list):
    lines = []
    audit_rows = []
    lines.append("=" * 78)
    lines.append("AUDIT — Rating statistico Serie A 2025/26 (target: FM 25/26)")
    lines.append("=" * 78)
    lines.append(
        f"\nJoin xlsx->scraped: {matched_count}/{sum(1 for x in QUOTAZIONI if x['FM'] and x['FM'] > 0)} "
        f"giocatori con FM>0 matchati.  Non matchati: {unmatched_list}"
    )

    for role in ("P", "D", "C", "A"):
        m = models[role]
        cv = m["cv"]
        lines.append(
            f"\n--- {ROLE_FANTASY[role]} ({role}) — n train={m['n_train']}, tot={m['n_total']} ---"
        )
        lines.append(
            f"  CV leave-one-out : R2={cv['r2']:.3f}  Pearson={cv['pearson']:.3f}  "
            f"Spearman={cv['spearman']:.3f}  RMSE={cv['rmse']:.3f} (FM)"
        )
        lines.append("  feature importance (effetto FM per +1 dev std della feature):")
        for feat, b in sorted(m["importance"], key=lambda t: -abs(t[1])):
            lines.append(f"    {feat[0]:22} beta={b:+.3f}")
        audit_rows.append(
            {
                "ruolo": role,
                "n_train": m["n_train"],
                "n_total": m["n_total"],
                "cv_r2": round(cv["r2"], 3),
                "cv_pearson": round(cv["pearson"], 3),
                "cv_spearman": round(cv["spearman"], 3),
                "cv_rmse": round(cv["rmse"], 3),
            }
        )
    lines.append("\n" + "=" * 78)

    # errori piu' forti per ruolo (FM reale vs predetto, dentro i matchati con Min>=270)
    lines.append(
        "\nSOTTOSTIMA / SOVRASTIMA piu' forti (FM - PredFM, solo con FM e Min>=270):"
    )
    for role in ("P", "D", "C", "A"):
        group = [
            r
            for r in ratings
            if r["role"] == role and r["has_fm"] and r["minutes"] >= MIN_MINUTES
        ]
        group.sort(key=lambda r: r["fm"] - r["pred"])
        worst = group[:3] + group[-3:]
        tag = ["sottostimati"] * 3 + ["sovrastimati"] * 3
        lines.append(f"  {ROLE_FANTASY[role]}:")
        for r, t in zip(worst, tag, strict=True):
            lines.append(
                f"    {t:12} {r['row']['Nome']:20} {r['row']['Squadra']:12} "
                f"FM={r['fm']:.2f} pred={r['pred']:.2f} d={r['fm'] - r['pred']:+.2f}"
            )

    # overlap Top-10 (FM reale vs rating) per ruolo
    lines.append("\nTOP-10 FM reali che sono anche TOP-10 del mio rating:")
    for role in ("P", "D", "C", "A"):
        group = [r for r in ratings if r["role"] == role and r["has_fm"]]
        by_fm = sorted(group, key=lambda r: -r["fm"])[:10]
        by_rating = sorted(group, key=lambda r: -r["rating"])[:10]
        fm_ids = {r["row"]["Id"] for r in by_fm}
        hit = sum(1 for r in by_rating if r["row"]["Id"] in fm_ids)
        lines.append(f"  {ROLE_FANTASY[role]}: {hit}/10")
        audit_rows[-1]["overlap_top10"] = f"{hit}/10"

    # copertura complessiva
    with_fm = sum(1 for r in ratings if r["has_fm"])
    lines.append(
        f"\nRating calcolati: {len(ratings)} giocatori (con FM per backtest: {with_fm})"
    )

    try:
        with open(OUT_AUDIT, "w", newline="", encoding="utf-8-sig") as f:
            cols = [
                "ruolo",
                "n_train",
                "n_total",
                "cv_r2",
                "cv_pearson",
                "cv_spearman",
                "cv_rmse",
                "overlap_top10",
            ]
            writer = csv.DictWriter(f, fieldnames=cols)
            writer.writeheader()
            writer.writerows(audit_rows)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_AUDIT}: {err}") from err

    print("\n".join(lines))
    try:
        with open(OUT_TXT, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_TXT}: {err}") from err


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #


def main():
    global SCRAPED, QUOTAZIONI, MATCHED, ROLE_OF
    SCRAPED = load_scraped()
    QUOTAZIONI = load_quotazioni()
    scraped_index = build_scraped_index(SCRAPED)

    # ruolo di ogni giocatore scrapato: quello dell'xlsx (autoritativo, con FM)
    # se matchato, altrimenti la roleLabel dell'API SDP
    ROLE_OF = {r["Id"]: SDP_TO_FANTASY.get(r["Ruolo"], "C") for r in SCRAPED}
    MATCHED = {}
    unmatched = []
    for x in QUOTAZIONI:
        if x["FM"] is None or x["FM"] <= 0:
            continue
        sp = match_player(
            x["Nome"], FANTASY_TO_SDP[x["R"]], x["Squadra"], scraped_index
        )
        if sp is None:
            unmatched.append(x["Nome"])
            continue
        pid = sp["row"]["Id"]
        MATCHED[pid] = x["FM"]
        ROLE_OF[pid] = x["R"]
    print(
        f"Giocatori scrapati: {len(SCRAPED)} | FM matchati: {len(MATCHED)} | persi: {len(unmatched)}"
    )

    models = {role: fit_role_model(role) for role in ("P", "D", "C", "A")}

    ratings = compute_ratings(models)
    out_rows = build_rows(ratings)
    write_csv(out_rows)
    write_xlsx(out_rows)
    run_audit(models, ratings, len(MATCHED), unmatched)

    print(
        f"\nOutput: {os.path.relpath(OUT_XLSX, BASE_DIR)}  /  {os.path.relpath(OUT_CSV, BASE_DIR)}"
    )
    print(f"Audit : {os.path.relpath(OUT_AUDIT, BASE_DIR)} (+ .txt)")


if __name__ == "__main__":
    main()
