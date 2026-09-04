#!/usr/bin/env -S uv run --quiet python
"""Extract official Fantacalcio Mantra roles into the runtime sidecar CSV."""

from __future__ import annotations

import argparse
import csv
import os
import tempfile
from contextlib import suppress
from pathlib import Path
from typing import Any, cast

import openpyxl
from mantra import parse_roles  # pyright: ignore[reportMissingImports]

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUT = BASE_DIR / "data" / "mantra_roles.csv"


def _header_row(sheet) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
        names = {
            str(value).strip(): i for i, value in enumerate(row) if value is not None
        }
        if {"RM", "Nome"}.issubset(names):
            return row_number, names
    raise ValueError("header con colonne RM e Nome non trovato")


def import_mantra_roles(source: str | Path, output: str | Path = DEFAULT_OUT) -> int:
    """Write ``nome,squadra,ruolo_mantra`` from an official quotations workbook."""
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = cast(
        Any,
        workbook["Tutti"] if "Tutti" in workbook.sheetnames else workbook.active,
    )
    if sheet is None:
        raise ValueError("il file non contiene un foglio dati Mantra")
    header_row, columns = _header_row(sheet)
    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[columns["Nome"]] or "").strip().upper()
        roles = parse_roles(row[columns["RM"]])
        if not name or not roles:
            continue
        if name in seen:
            raise ValueError(f"nome duplicato nel listone Mantra: {name}")
        seen.add(name)
        team_index = columns.get("Squadra")
        team = str(row[team_index] or "").strip() if team_index is not None else ""
        records.append({"nome": name, "squadra": team, "ruolo_mantra": ";".join(roles)})
    if not records:
        raise ValueError("nessun ruolo Mantra valido trovato")
    records.sort(key=lambda record: record["nome"])
    destination = Path(output)
    destination.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(
        prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent
    )
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(
                stream, fieldnames=["nome", "squadra", "ruolo_mantra"]
            )
            writer.writeheader()
            writer.writerows(records)
        os.replace(temporary, destination)
    except BaseException:
        with suppress(FileNotFoundError):
            os.unlink(temporary)
        raise
    return len(records)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", help="file Quotazioni Fantacalcio .xlsx")
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()
    count = import_mantra_roles(args.source, args.out)
    print(f"Importati {count} ruoli Mantra in {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
