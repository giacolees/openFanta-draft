#!/usr/bin/env -S uv run --quiet python
"""
Player Rating Stagione 2026/27 — Fantacalcio.

Costruisce un rating stagionale (0-100) per confrontare i giocatori a parita' di ruolo,
basato esclusivamente su data/Quotazioni_Fantacalcio_Stagione_2026_27_arricchito.xlsx.

Metodo:
- Percentile di ogni metrica 25/26 calcolato DENTRO il ruolo (P, D, C, A).
- Media pesata con pesi specifici per ruolo (fantamedia come segnale principale,
  produzione attaccante/difensiva, affidabilita' sulle presenze, contesto di mercato FVM).
- Shrinkage sull'affidabilita': chi ha poche presenze viene riportato verso la mediana del ruolo.
- Giocatori senza statistiche 25/26 (nuovi acquisti): rating provvisorio dal percentile FVM.

Output:
- data/Player_Rating_Stagione_2026_27.xlsx (sheet Tutti + un sheet per ruolo)
- data/player_rating_stagione_2026_27.csv
"""

import csv
import os
from collections import defaultdict

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_XLSX = os.path.join(
    BASE_DIR, "data", "Quotazioni_Fantacalcio_Stagione_2026_27_arricchito.xlsx"
)
OUT_XLSX = os.path.join(BASE_DIR, "data", "Player_Rating_Stagione_2026_27.xlsx")
OUT_CSV = os.path.join(BASE_DIR, "data", "player_rating_stagione_2026_27.csv")

ROLE_SHEETS = {
    "Portieri": "P",
    "Difensori": "D",
    "Centrocampisti": "C",
    "Attaccanti": "A",
}
ROLE_LABEL = {
    "P": "Portiere",
    "D": "Difensore",
    "C": "Centrocampista",
    "A": "Attaccante",
}

# Pesi per ruolo: (colonna 25/26, peso, modalita')
# modalita': "raw" = valore cosi' com'e' | "per90" = ogni 90 minuti (richiede Min >= 270)
#            | "ratio_pres" = diviso le presenze (richiede Pres >= 5)
ROLE_WEIGHTS = {
    "P": [
        ("FM 25/26", 0.40, "raw"),
        ("MV 25/26", 0.10, "raw"),
        ("CS 25/26", 0.20, "ratio_pres"),
        ("Parate 25/26", 0.15, "per90"),
        ("Pres 25/26", 0.10, "raw"),
        ("FVM", 0.05, "raw"),
    ],
    "D": [
        ("FM 25/26", 0.35, "raw"),
        ("MV 25/26", 0.10, "raw"),
        ("Gol 25/26", 0.10, "per90"),
        ("Assist 25/26", 0.10, "per90"),
        ("CS 25/26", 0.15, "ratio_pres"),
        ("xA 25/26", 0.05, "per90"),
        ("Pres 25/26", 0.10, "raw"),
        ("FVM", 0.05, "raw"),
    ],
    "C": [
        ("FM 25/26", 0.35, "raw"),
        ("MV 25/26", 0.10, "raw"),
        ("Gol 25/26", 0.15, "per90"),
        ("Assist 25/26", 0.15, "per90"),
        ("GA 25/26", 0.10, "per90"),
        ("Pres 25/26", 0.10, "raw"),
        ("FVM", 0.05, "raw"),
    ],
    "A": [
        ("FM 25/26", 0.35, "raw"),
        ("MV 25/26", 0.10, "raw"),
        ("Gol 25/26", 0.20, "per90"),
        ("Assist 25/26", 0.05, "per90"),
        ("xG 25/26", 0.10, "per90"),
        ("TiriPorta 25/26", 0.05, "per90"),
        ("Pres 25/26", 0.10, "raw"),
        ("FVM", 0.05, "raw"),
    ],
}

KEEP_COLS = [
    "Id",
    "R",
    "RM",
    "Nome",
    "Squadra",
    "Qt.A",
    "FVM",
    "Pres 25/26",
    "Min 25/26",
    "MV 25/26",
    "FM 25/26",
    "Gol 25/26",
    "Assist 25/26",
    "CS 25/26",
    "Parate 25/26",
    "xG 25/26",
    "xA 25/26",
    "GA 25/26",
    "TiriPorta 25/26",
]


def to_float(value):
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(
            value if isinstance(value, (int, float)) else str(value).replace(",", ".")
        )
    except (TypeError, ValueError):
        return None


def to_int(value):
    numeric = to_float(value)
    if numeric is None:
        return None
    try:
        return int(numeric)
    except (OverflowError, TypeError, ValueError):
        return None


def load_players():
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    players = []
    for sheet, role in ROLE_SHEETS.items():
        ws = wb[sheet]
        if not isinstance(ws, Worksheet):
            raise TypeError(f"Il foglio {sheet!r} non e' un worksheet dati")
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else "" for h in rows[1]]
        idx = {name: header.index(name) for name in KEEP_COLS if name in header}
        for row in rows[2:]:
            if row is None or idx.get("Nome") is None or row[idx["Nome"]] in (None, ""):
                continue
            p = {name: row[i] if i < len(row) else None for name, i in idx.items()}
            for name in KEEP_COLS:
                if name not in ("Nome", "Squadra", "RM"):
                    p[name] = to_float(p.get(name))
            p["R"] = role
            p["Nome"] = str(p["Nome"]).strip()
            p["Squadra"] = str(p.get("Squadra") or "").strip()
            p["RM"] = str(p.get("RM") or "").strip()
            players.append(p)
    return players


def derived_metrics(p):
    """Aggiunge le metriche derivate usate nella valutazione."""
    pres = p.get("Pres 25/26")
    minutes = p.get("Min 25/26")
    per90_base = minutes if minutes and minutes > 0 else (pres * 90 if pres else None)
    for stat in ("Gol", "Assist", "Parate", "xG", "xA", "GA", "TiriPorta"):
        col = f"{stat} 25/26"
        value = p.get(col)
        if per90_base and per90_base >= 270 and value is not None:
            p[col + "/90"] = value * 90 / per90_base
        else:
            p[col + "/90"] = None
    if pres and pres >= 5 and p.get("CS 25/26") is not None:
        p["CS%"] = p["CS 25/26"] / pres
    else:
        p["CS%"] = None
    return p


def percentile_ranks(players):
    """Percentile (0-1) di ogni metrica numerica, calcolato dentro il ruolo."""
    by_role = defaultdict(list)
    for p in players:
        by_role[p["R"]].append(p)
    ranks = {}  # (id_giocatore, metrica) -> percentile
    for group in by_role.values():
        metrics = set()
        for p in group:
            metrics.update(k for k, v in p.items() if isinstance(v, float))
        for metric in metrics:
            values = sorted(p[metric] for p in group if p.get(metric) is not None)
            n = len(values)
            if n < 2:
                continue
            for p in group:
                v = p.get(metric)
                if v is None:
                    continue
                below = sum(1 for x in values if x < v)
                equal = sum(1 for x in values if x == v)
                ranks[(id(p), metric)] = (below + equal / 2) / n
    return ranks


def composite_rating(p, ranks, role_ranks):
    """Rating composito 0-100 per un giocatore."""
    has_season = any(
        p.get(c) is not None for c in ("FM 25/26", "MV 25/26", "Pres 25/26")
    )
    if not has_season:
        # Nuovo acquisto: stima dal mercato (percentile FVM) ma scontata,
        # perche' senza minutaggio in Serie A non puo' superare chi ha rendimento provato.
        fvm_rank = role_ranks["FVM"].get(id(p))
        if fvm_rank is None:
            return 50.0, "Nuovo (stima su FVM)"
        shrunk = 0.6 * fvm_rank + 0.4 * 0.5
        return 100 * shrunk, "Nuovo (stima su FVM)"

    weights = ROLE_WEIGHTS[p["R"]]
    num = den = 0.0
    for col, weight, mode in weights:
        key = col + (
            "/90"
            if mode == "per90"
            else ("%" if mode == "ratio_pres" and col == "CS 25/26" else "")
        )
        rank = (
            role_ranks[key].get(id(p)) if key in role_ranks else ranks.get((id(p), key))
        )
        if rank is None:
            continue
        num += weight * rank
        den += weight
    base = num / den if den else 0.5

    pres = p.get("Pres 25/26") or 0
    reliability = min(
        pres / 25.0, 1.0
    )  # sotto le 25 presenze si riporta verso la mediana
    shrunk = reliability * base + (1 - reliability) * 0.5
    return 100 * shrunk, "25/26"


def tier(rank_fraction):
    if rank_fraction >= 0.90:
        return "S"
    if rank_fraction >= 0.70:
        return "A"
    if rank_fraction >= 0.30:
        return "B"
    if rank_fraction >= 0.10:
        return "C"
    return "D"


OUT_COLUMNS = [
    "Rating",
    "Tier",
    "Rank ruolo",
    "Ruolo",
    "RM",
    "Nome",
    "Squadra",
    "Id",
    "Qt.A",
    "FVM",
    "Pres",
    "Min",
    "MV",
    "FM",
    "Gol",
    "Assist",
    "CS",
    "Parate",
    "xG",
    "xA",
    "GA",
    "TiriPorta",
    "Dati",
]

SOURCE_MAP = [
    ("Pres 25/26", "Pres"),
    ("Min 25/26", "Min"),
    ("MV 25/26", "MV"),
    ("FM 25/26", "FM"),
    ("Gol 25/26", "Gol"),
    ("Assist 25/26", "Assist"),
    ("CS 25/26", "CS"),
    ("Parate 25/26", "Parate"),
    ("xG 25/26", "xG"),
    ("xA 25/26", "xA"),
    ("GA 25/26", "GA"),
    ("TiriPorta 25/26", "TiriPorta"),
]


def build_rows(players, ranks, role_ranks):
    for p in players:
        p["Rating"], p["Dati"] = composite_rating(p, ranks, role_ranks)
    by_role = defaultdict(list)
    for p in players:
        by_role[p["R"]].append(p)

    rows = []
    for role, group in by_role.items():
        group.sort(key=lambda p: p["Rating"], reverse=True)
        n = len(group)
        for i, p in enumerate(group, start=1):
            row = {
                "Rating": round(p["Rating"], 1),
                "Tier": tier(1 - (i - 1) / max(n - 1, 1)),
                "Rank ruolo": i,
                "Ruolo": ROLE_LABEL[role],
                "RM": p["RM"],
                "Nome": p["Nome"],
                "Squadra": p["Squadra"],
                "Id": to_int(p.get("Id")),
                "Qt.A": p.get("Qt.A"),
                "FVM": p.get("FVM"),
                "Dati": p["Dati"],
            }
            for src, dst in SOURCE_MAP:
                if dst in ("Pres", "Min"):
                    v = p.get(src)
                    row[dst] = to_int(v)
                elif dst in ("MV", "FM"):
                    v = p.get(src)
                    row[dst] = round(v, 2) if v is not None else None
                elif dst == "Gol" or dst == "Assist" or dst == "CS" or dst == "Parate":
                    v = p.get(src)
                    row[dst] = to_int(v)
                else:
                    v = p.get(src)
                    row[dst] = round(v, 2) if v is not None else None
            rows.append(row)
    rows.sort(key=lambda r: (r["Ruolo"], r["Rank ruolo"]))
    return rows


def write_csv(rows):
    try:
        with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=OUT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
    except OSError as err:
        raise RuntimeError(f"Impossibile scrivere {OUT_CSV}: {err}") from err


def write_xlsx(rows):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor=Color(rgb="1F4E78"))
    header_font = Font(bold=True, color=Color(rgb="FFFFFF"))
    sheets = [("Tutti", rows)] + [
        (ruolo, [r for r in rows if r["Ruolo"] == label])
        for ruolo, label in [
            ("Portieri", "Portiere"),
            ("Difensori", "Difensore"),
            ("Centrocampisti", "Centrocampista"),
            ("Attaccanti", "Attaccante"),
        ]
    ]
    for name, sheet_rows in sheets:
        ws = wb.create_sheet(name)
        if not isinstance(ws, Worksheet):
            raise TypeError(f"Impossibile creare il worksheet dati {name!r}")
        ws.append(OUT_COLUMNS)
        for header_row in ws.iter_rows(min_row=1, max_row=1):
            for cell in header_row:
                if not isinstance(cell, Cell):
                    continue
                cell.fill, cell.font = header_fill, header_font
                cell.alignment = Alignment(horizontal="center")
        for r in sheet_rows:
            ws.append([r[c] for c in OUT_COLUMNS])
        for col_idx, col in enumerate(OUT_COLUMNS, start=1):
            width = max(len(col) + 2, 8) if col != "Nome" else 20
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "H2"
    del wb["Sheet"]
    wb.save(OUT_XLSX)


def main():
    players = [derived_metrics(p) for p in load_players()]
    ranks = percentile_ranks(players)

    # indici rapidi per percentile FVM e chiavi composite, dentro il ruolo
    role_ranks = defaultdict(dict)
    by_role = defaultdict(list)
    for p in players:
        by_role[p["R"]].append(p)
    for role, group in by_role.items():
        fvms = sorted(p["FVM"] for p in group if p.get("FVM") is not None)
        for p in group:
            v = p.get("FVM")
            if v is not None and len(fvms) > 1:
                below = sum(1 for x in fvms if x < v)
                equal = sum(1 for x in fvms if x == v)
                role_ranks["FVM"][id(p)] = (below + equal / 2) / len(fvms)
        for col, _, mode in ROLE_WEIGHTS[role]:
            key = col + (
                "/90"
                if mode == "per90"
                else ("%" if mode == "ratio_pres" and col == "CS 25/26" else "")
            )
            role_ranks[key] = {pid: r for (pid, m), r in ranks.items() if m == key}

    rows = build_rows(players, ranks, role_ranks)
    write_csv(rows)
    write_xlsx(rows)

    print(f"Giocatori valutati: {len(rows)}  ->  {os.path.relpath(OUT_XLSX, BASE_DIR)}")
    print(
        f"                                  ->  {os.path.relpath(OUT_CSV, BASE_DIR)}\n"
    )
    for label in ("Portiere", "Difensore", "Centrocampista", "Attaccante"):
        top = [r for r in rows if r["Ruolo"] == label][:10]
        print(f"=== TOP 10 {label.upper()} ===")
        for r in top:
            print(
                f"  {r['Rank ruolo']:>3}. {r['Nome']:<22} {r['Squadra']:<12} "
                f"Rating {r['Rating']:>5}  Tier {r['Tier']}  Qt.A {r['Qt.A']}"
            )
        print()


if __name__ == "__main__":
    main()
