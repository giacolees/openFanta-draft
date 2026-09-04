#!/usr/bin/env -S uv run --quiet python
"""
Importatore del Listone Fantaculo — Fantacalcio Stagione 2026/27 (WP3).

Legge il file Excel del listone (il nome cambia con la data: Listone_Fantaculo_AAAA_MM_GG.xlsx,
di default prende il piu' recente in data/) e produce il CSV canonico data/listone.csv
consumato da asta CLI e GUI, piu' i sidecar:
- data/listone_meta.json      : stagione, fonte, data import, conteggi/totali PFC per ruolo,
                                colonne, algoritmo/versione del pid.
- data/import_report_<ts>.json: errori/warning dell'import (strutturali, dati, duplicati/collisioni).

## Identita' del giocatore — pid
Il foglio non ha un ID sorgente: il pid e' un hash deterministico VERSIONATO di
``norm(nome) | norm(ruolo)`` (FNV-1a 64-bit, esadecimale a 16 cifre). La squadra
NON entra nel pid: l'identita' deve sopravvivere ai trasferimenti tra i refresh
giornalieri del listone. Collisione o duplicato di pid = errore BLOCANTE, mai
unione automatica; per omonimi reali nello stesso ruolo (stesso nome normalizzato
e ruolo ma squadra diversa) il pid collide e l'import si blocca riportandoli (non
si inventano suffissi). Un cambio di algoritmo/input del hash => bump di
``PID_VERSION`` (i pid cambiano: documentato nei meta, e' voluto).

Colonne chiave esportate:
- pid            : identita' stabile (FNV-1a 64 hex) — vedi sopra
- pfc, pma       : prezzi suggeriti Fantaculo e medi d'Italia
- pfc/pma_range  : range originali "min-max" (stringhe preservate)
- pfc/pma_lo,hi  : estremi numerici parse dei range (None se non parsabili)
- unc_pfc        : incertezza = semiampiezza del range PFC = round((hi-lo)/2, 1)
- slot (1..8)    : fascia di qualita' nel ruolo (1 = top del ruolo ... 8 = coda)
- tit 0-100      : expected titolarita'
- expfm 0-10     : expected fantamedia per partita
- tix / fix      : percentile nel ruolo di titolarita' / del contributo bonus
- fix_contrib    : contributo bonus atteso = (expfm - 6) * tit/100

## Validazione e uscite
- Header obbligatori (name, team, role, pfc, pma, slot, expectedTitolarita,
  expectedFantamedia, pfcRange, pmaRange): mancanti => errore BLOCANTE.
- Dati: pfc numero > 0 BLOCANTE; slot intero 1..8, tit 0..100, expfm 0..10 BLOCANTI;
  pma numero > 0 -> se mancante/non positivo WARNING e stima = pfc (3 giocatori reali
  hanno pma=0 e sono recuperabili: la media d'Italia non e' mai obbligatoria).
- Duplicati sul pid (stesso nome normalizzato + ruolo) => BLOCANTE, riportati in
  ``pid_collisions`` (mai unione automatica).
- Errori BLOCANTI => exit code 2, NESSUN sovrascrittura del CSV/meta validi
  (scrittura atomica su file temporaneo + rename). Solo warning => exit 0.
- Sempre scritto il report JSON ``data/import_report_<ts>.json``.

Uso:
  uv run scripts/import_listone.py                 # ultimo listone in data/
  uv run scripts/import_listone.py --file PATH     # file specifico
  uv run scripts/import_listone.py --out PATH      # CSV di destinazione
"""

import argparse
import csv
import glob
import json
import os
import re
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any

import openpyxl
from league_config import ROLE_ORDER, norm
from mantra import parse_roles  # pyright: ignore[reportMissingImports]
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
PATTERN = os.path.join(DATA_DIR, "Listone_Fantaculo_*.xlsx")
DEFAULT_OUT = os.path.join(DATA_DIR, "listone.csv")
META_PATH = os.path.join(DATA_DIR, "listone_meta.json")

SEASON = "2026-27"

# ----------------------------------------------------------- identita' (pid)
# Algoritmo/versione documentati anche in listone_meta.json. Un bump di
# PID_VERSION significa "il modo di derivare il pid e' cambiato": i pid NEI
# CSV nuovi saranno diversi da quelli vecchi (voluto). L'input del hash NON
# include la squadra: il pid sopravvive ai trasferimenti tra refresh.
PID_ALGORITHM = "fnv1a64(norm(nome) + '|' + norm(ruolo))"
PID_VERSION = "1"

_FNV1A64_OFFSET = 0xCBF29CE484222325
_FNV1A64_PRIME = 0x00000100000001B3
_PID_RE = re.compile(r"^[0-9a-f]{16}$")


def fnv1a64(data: str) -> str:
    """FNV-1a a 64 bit, esadecimale a 16 cifre (deterministico, collisione ~2^-64)."""
    h = _FNV1A64_OFFSET
    for byte in data.encode("utf-8"):
        h ^= byte
        h = (h * _FNV1A64_PRIME) & 0xFFFFFFFFFFFFFFFF
    return f"{h:016x}"


def compute_pid(nome: str, ruolo: str) -> str:
    """PID deterministica e stabile del giocatore: hash di norm(nome)|norm(ruolo).

    La squadra e' ESPLICITAMENTE esclusa: il pid deve sopravvivere ai trasferimenti
    tra i refresh giornalieri del listone.
    """
    return fnv1a64(f"{norm(nome)}|{norm(ruolo)}")


def is_pid(text: str) -> bool:
    return bool(_PID_RE.match(text))


# --------------------------------------------------------------- range numerici
def parse_range(text) -> tuple[float, float] | None:
    """Parsa una stringa 'min-max' -> (lo, hi) float, con lo <= hi (ordina se invertito).

    Ritorna None se la stringa e' assente o non parsabile (range sporco): in quel
    caso lo/hi non vengono inventati, restano None e si emette un warning.
    Accetta separatori - / – / — / ~ e spazi, e virgole decimali.
    """
    if text is None:
        return None
    s = str(text).replace(",", ".").strip()
    if not s:
        return None
    m = re.match(r"^\s*(-?\d+(?:\.\d+)?)\s*[-–—~]\s*(-?\d+(?:\.\d+)?)\s*$", s)
    if not m:
        return None
    try:
        lo, hi = float(m.group(1)), float(m.group(2))
    except (TypeError, ValueError):
        return None
    if lo > hi:
        lo, hi = hi, lo
    return (lo, hi)


# ------------------------------------------------------------------- schema CSV
CANON = [
    "pid",
    "nome",
    "squadra",
    "ruolo",
    "ruolo_mantra",
    "pfc",
    "pma",
    "pfc_range",
    "pma_range",
    "pfc_lo",
    "pfc_hi",
    "pma_lo",
    "pma_hi",
    "unc_pfc",
    "dpfcpma",
    "slot",
    "tit",
    "expfm",
    "fascia",
    "status",
    "pen_prob",
    "fk_prob",
    "tix",
    "fix",
    "fix_contrib",
]

# colonne del foglio ALL -> nomi canonici (il listone tronca alcune intestazioni a 22 caratteri)
COLUMN_MAP = {
    "name": "nome",
    "team": "squadra",
    "role": "ruolo",
    "roleMantra": "ruolo_mantra",
    "pfc": "pfc",
    "pma": "pma",
    "dpfcpma": "dpfcpma",
    "pfcRange": "pfc_range",
    "pmaRange": "pma_range",
    "slot": "slot",
    "expectedTitolarita": "tit",
    "expectedFantamedia": "expfm",
    "fasciaFc": "fascia",
    "playerStatus": "status",
    "penaltyProbability": "pen_prob",
    "freeKickProbability": "fk_prob",
}

# colonne del foglio obbligatorie: se manca una di queste, l'import si blocca.
REQUIRED_SOURCE_COLUMNS = {
    "name",
    "team",
    "role",
    "pfc",
    "pma",
    "pfcRange",
    "pmaRange",
    "slot",
    "expectedTitolarita",
    "expectedFantamedia",
}


def latest_file(pattern=PATTERN):
    files = glob.glob(pattern)
    if not files:
        raise SystemExit(f"Nessun listone trovato: {pattern}")

    def sort_key(path):
        m = re.search(r"(\d{4})[_-](\d{2})[_-](\d{2})", os.path.basename(path))
        return (m.groups() if m else "0000-00-00", os.path.getmtime(path))

    return max(files, key=sort_key)


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def to_int(value, default=None):
    v = to_float(value)
    if v is None:
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def percentile_ranks(rows, key):
    """Percentile 0-100 di rows[key] calcolato dentro il ruolo."""
    by_role = defaultdict(list)
    for r in rows:
        if r.get(key) is not None:
            by_role[r["ruolo"]].append(r[key])
    ranks = {}
    for role, values in by_role.items():
        s = sorted(values)
        n = len(s)
        if n < 2:
            ranks[role] = dict.fromkeys(set(values), 50.0)
            continue
        out = {}
        for v in set(values):
            below = sum(1 for x in s if x < v)
            equal = sum(1 for x in s if x == v)
            out[v] = 100 * (below + equal / 2) / n
        ranks[role] = out
    return ranks


# ---------------------------------------------------------------------- import
def import_listone(path, out_path=DEFAULT_OUT, meta_path=META_PATH):
    """Importa il listone xlsx -> CSV canonico + meta + report.

    Ritorna un dict strutturato (per test e CLI): ``ok`` (no errori bloccanti),
    ``errors``, ``warnings``, ``n_imported``, ``n_skipped``, ``pid_collisions``,
    ``players`` (le righe validate; vuoto se bloccato), ``meta``.

    Su errori bloccanti NON scrive il CSV ne' il meta (l'ultimo valido resta
    intatto: scrittura atomica); scrive sempre il report JSON. L'eventuale CSV
    viene scritto atomicamente (temp + rename) in ogni caso di successo.
    """
    errors: list[str] = []
    warnings: list[str] = []
    n_skipped = 0

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = _pick_data_sheet(wb)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        errors.append("foglio vuoto: nessuna riga di dati")
        return _result(False, errors, warnings, [], [], {}, out_path, n_skipped)

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    src_to_canon = dict(COLUMN_MAP)
    missing_headers = sorted(REQUIRED_SOURCE_COLUMNS - set(header))
    if missing_headers:
        errors.append("header obbligatori mancanti: " + ", ".join(missing_headers))
        # mancando colonne strutturali non si puo' procedere: report e basta
        return _result(False, errors, warnings, [], [], {}, out_path, n_skipped)

    players = []
    for linenum, raw in enumerate(rows[1:], start=2):
        if raw is None or raw[0] is None or not str(raw[0]).strip():
            n_skipped += 1
            continue
        rec = dict(zip(header, raw, strict=False))
        p = {canon: rec.get(src) for src, canon in src_to_canon.items()}
        p["nome"] = str(p["nome"]).strip().upper()
        p["squadra"] = str(p.get("squadra") or "").strip()
        p["ruolo"] = str(p.get("ruolo") or "").strip().upper()
        p["ruolo_mantra"] = ";".join(parse_roles(p.get("ruolo_mantra")))
        ctx = f"riga {linenum} ({p['nome']}, {p['ruolo']})"

        if p["ruolo"] not in ROLE_ORDER:
            n_skipped += 1
            continue

        # ---- tipi numerici con validazione dei range -----------------------
        pfc = to_float(p.get("pfc"))
        if pfc is None or pfc <= 0:
            errors.append(
                f"{ctx}: pfc non valido o non positivo (trovato {p.get('pfc')!r})"
            )
            continue
        p["pfc"] = pfc

        pma = to_float(p.get("pma"))
        if pma is None or pma <= 0:
            # pma non e' mai obbligatorio (media d'Italia; 3 giocatori reali hanno 0):
            # warning + stima = pfc, non dati inventati ma fallback documentato.
            warnings.append(
                f"{ctx}: pma non positivo o non valido (trovato {p.get('pma')!r}), "
                f"usato il pfc come stima"
            )
            pma = pfc
        p["pma"] = pma

        slot = to_int(p.get("slot"))
        if slot is None or not (1 <= slot <= 8):
            errors.append(
                f"{ctx}: slot non valido (trovato {p.get('slot')!r}, atteso 1-8)"
            )
            continue
        try:
            p["slot"] = int(slot)
        except (TypeError, ValueError):
            errors.append(f"{ctx}: slot non intero (trovato {p.get('slot')!r})")
            continue

        tit = to_float(p.get("tit"))
        if tit is None or not (0 <= tit <= 100):
            errors.append(
                f"{ctx}: titolarita' non valida (trovato {p.get('tit')!r}, attesa 0-100)"
            )
            continue
        p["tit"] = tit

        expfm = to_float(p.get("expfm"))
        if expfm is None or not (0 <= expfm <= 10):
            errors.append(
                f"{ctx}: fantamedia attesa non valida (trovato {p.get('expfm')!r}, attesa 0-10)"
            )
            continue
        p["expfm"] = expfm

        dpfcpma = to_float(p.get("dpfcpma"))
        p["dpfcpma"] = dpfcpma if dpfcpma is not None else round(pfc - pma, 1)
        p["pen_prob"] = to_float(p.get("pen_prob"))
        p["fk_prob"] = to_float(p.get("fk_prob"))

        # ---- range numerici dai range stringa (preservati come stringhe) ----
        p["pfc_range"] = str(p.get("pfc_range") or "").strip()
        p["pma_range"] = str(p.get("pma_range") or "").strip()
        pfc_r = parse_range(p["pfc_range"])
        pma_r = parse_range(p["pma_range"])
        if p["pfc_range"] and pfc_r is None:
            warnings.append(
                f"{ctx}: range PFC non parsabile '{p['pfc_range']}' -> pfc_lo/hi=None"
            )
        if p["pma_range"] and pma_r is None:
            warnings.append(
                f"{ctx}: range PMA non parsabile '{p['pma_range']}' -> pma_lo/hi=None"
            )
        p["pfc_lo"], p["pfc_hi"] = pfc_r if pfc_r else (None, None)
        p["pma_lo"], p["pma_hi"] = pma_r if pma_r else (None, None)
        p["unc_pfc"] = (
            round((p["pfc_hi"] - p["pfc_lo"]) / 2, 1)
            if p["pfc_lo"] is not None and p["pfc_hi"] is not None
            else None
        )

        p["fascia"] = str(p.get("fascia") or "").strip()
        p["status"] = str(p.get("status") or "").strip()

        # ---- identita': pid ------------------------------------------------
        p["pid"] = compute_pid(p["nome"], p["ruolo"])

        # contributo bonus atteso a giornata: il voto base 6 e' gratis (FM-6 per titolarita')
        p["fix_contrib"] = round((expfm - 6.0) * tit / 100.0, 3)
        players.append(p)

    # ---- duplicati / collisioni di identita' (bloccanti, mai merge) --------
    pid_collisions: list[dict[str, Any]] = []
    seen_nsr: dict[tuple, str] = {}
    pid_index: dict[str, list[dict]] = defaultdict(list)
    for p in players:
        pid_index[p["pid"]].append(p)
        key = (norm(p["nome"]), norm(p["squadra"]), p["ruolo"])
        if key in seen_nsr:
            errors.append(
                f"riga duplicata (stesso nome+squadra+ruolo di '{seen_nsr[key]}'): "
                f"{p['nome']}, {p['squadra'] or '—'}, {p['ruolo']}"
            )
        seen_nsr[key] = p["nome"]
    for pid, group in pid_index.items():
        if len(group) > 1:
            pid_collisions.append(
                {
                    "pid": pid,
                    "nome": group[0]["nome"],
                    "ruolo": group[0]["ruolo"],
                    "players": [
                        {
                            "nome": p["nome"],
                            "squadra": p["squadra"],
                            "ruolo": p["ruolo"],
                        }
                        for p in group
                    ],
                }
            )
            errors.append(
                f"collisione di identita': pid {pid} per {len(group)} giocatori "
                f"(stesso nome normalizzato '{norm(group[0]['nome'])}' e ruolo "
                f"{group[0]['ruolo']}) — omonimi reali o duplicati, non unibili "
                f"automaticamente: {', '.join(p['nome'] for p in group)}"
            )

    ok = not errors
    players_out = [] if not ok else players

    if ok:
        # ---- indici percentile nel ruolo (tix / fix) -----------------------
        tit_ranks = percentile_ranks(players, "tit")
        fix_ranks = percentile_ranks(players, "fix_contrib")
        for p in players:
            p["tix"] = round(tit_ranks[p["ruolo"]].get(p["tit"], 50.0), 1)
            p["fix"] = round(fix_ranks[p["ruolo"]].get(p["fix_contrib"], 50.0), 1)

        players.sort(key=lambda r: (r["ruolo"], -r["pfc"]))
        _atomic_write_csv(out_path, players)

    meta = _build_meta(path, players_out if ok else [], ok, warnings)
    if ok:
        _atomic_write_json(meta_path, meta)

    return _result(
        ok, errors, warnings, players_out, pid_collisions, meta, out_path, n_skipped
    )


def _result(ok, errors, warnings, players, pid_collisions, meta, out_path, n_skipped):
    return {
        "ok": ok,
        "errors": list(errors),
        "warnings": list(warnings),
        "n_imported": len(players) if ok else 0,
        "n_skipped": n_skipped,
        "pid_collisions": pid_collisions,
        "players": list(players),
        "meta": meta,
        "out_path": out_path,
    }


def _pick_data_sheet(wb) -> Worksheet:
    """Foglio dei dati: 'ALL' se presente, altrimenti il primo foglio Worksheet.
    Filtra via Chartsheet/WriteOnlyWorksheet (che non hanno iter_rows)."""
    names = list(wb.sheetnames)
    if "ALL" in names:
        return wb["ALL"]
    for name in names:
        cand = wb[name]
        if isinstance(cand, Worksheet):
            return cand
    raise ValueError("nessun foglio di dati nel workbook")


def _build_meta(path, players, ok, warnings):
    n_by_role = dict.fromkeys(ROLE_ORDER, 0)
    total_pfc_by_role = dict.fromkeys(ROLE_ORDER, 0.0)
    for p in players:
        n_by_role[p["ruolo"]] += 1
        total_pfc_by_role[p["ruolo"]] += p["pfc"]
    return {
        "season": SEASON,
        "source_file": os.path.basename(path),
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "n_players": len(players) if ok else 0,
        "n_by_role": n_by_role,
        "total_pfc_by_role": {k: round(v, 1) for k, v in total_pfc_by_role.items()},
        "total_pfc": round(sum(total_pfc_by_role.values()), 1),
        "columns": list(CANON),
        "pid_algorithm": PID_ALGORITHM,
        "pid_version": PID_VERSION,
        "ok": ok,
        "warnings": list(warnings),
    }


def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except OSError as e:
        raise OSError(f"impossibile creare la directory {path}: {e}") from e


def _atomic_write_csv(out_path, players):
    """Scrittura atomica del CSV: temp nella stessa dir + rename (mai CSV tronco)."""
    _ensure_dir(os.path.dirname(os.path.abspath(out_path)))
    fd, tmp = tempfile.mkstemp(
        dir=os.path.dirname(os.path.abspath(out_path)), suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=CANON, restval="", extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(players)
        os.replace(tmp, out_path)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise


def _atomic_write_json(path, obj):
    _ensure_dir(os.path.dirname(os.path.abspath(path)))
    fd, tmp = tempfile.mkstemp(
        dir=os.path.dirname(os.path.abspath(path)), suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(tmp, path)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise


def print_summary(players, path, src_name, warnings=()):
    print(f"Listone: {os.path.basename(src_name)}")
    print(f"Giocatori importati: {len(players)}  ->  {os.path.relpath(path, BASE_DIR)}")
    for role in ROLE_ORDER:
        g = [p for p in players if p["ruolo"] == role]
        if not g:
            continue
        tot = sum(p["pfc"] for p in g)
        print(
            f"  {role}: {len(g):>3} giocatori, Somma PFC {tot:>7.0f} cr, "
            f"PFC medio {tot / len(g):>5.1f}"
        )


def main():
    ap = argparse.ArgumentParser(description="Importatore Listone Fantaculo (WP3)")
    ap.add_argument(
        "--file", help="file xlsx specifico (default: il piu' recente in data/)"
    )
    ap.add_argument("--out", default=DEFAULT_OUT, help="CSV canonico di destinazione")
    ap.add_argument(
        "--report",
        default=None,
        help="report JSON (default: data/import_report_<ts>.json)",
    )
    ap.add_argument(
        "--meta", default=None, help="meta JSON (default: data/listone_meta.json)"
    )
    args = ap.parse_args()

    src = args.file or latest_file()
    meta_path = args.meta or META_PATH
    report = import_listone(src, args.out, meta_path)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    report_path = args.report or os.path.join(DATA_DIR, f"import_report_{ts}.json")
    _atomic_write_json(
        report_path,
        {
            "source_file": os.path.basename(src),
            "season": SEASON,
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "ok": report["ok"],
            "errors": report["errors"],
            "warnings": report["warnings"],
            "n_imported": report["n_imported"],
            "n_skipped": report["n_skipped"],
            "pid_collisions": report["pid_collisions"],
        },
    )

    if report["errors"]:
        for e in report["errors"]:
            print(f"Errore: {e}")
        print(
            f"Import BLOCCATO: {len(report['errors'])} errori (l'ultimo CSV valido resta intatto)."
        )
        print(f"Report: {os.path.relpath(report_path, BASE_DIR)}")
        raise SystemExit(2)
    for w in report["warnings"]:
        print(f"Avviso: {w}")
    print_summary(report["players"], report["out_path"], src, report["warnings"])
    if report["warnings"]:
        print(f"Warning: {len(report['warnings'])} (non bloccanti)")
    print(f"Report: {os.path.relpath(report_path, BASE_DIR)}")
    print(f"Meta:   {os.path.relpath(meta_path, BASE_DIR)}")


if __name__ == "__main__":
    main()
